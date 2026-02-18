import pandas as pd
import numpy as np
import re
from datetime import date
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import text, delete
from typing import Tuple, Dict, Any, Optional
from app.models_star_schema import (
    DimTime, DimRegion, DimProduct, FactSales, FactProductPerformance,
    get_month_name, get_month_short, get_quarter, get_fiscal_year
)


# Month name to number mapping
MONTH_MAP = {
    'january': 1, 'jan': 1,
    'february': 2, 'feb': 2,
    'march': 3, 'mar': 3,
    'april': 4, 'apr': 4,
    'may': 5,
    'june': 6, 'jun': 6,
    'july': 7, 'jul': 7,
    'august': 8, 'aug': 8,
    'september': 9, 'sep': 9, 'sept': 9,
    'october': 10, 'oct': 10,
    'november': 11, 'nov': 11,
    'december': 12, 'dec': 12
}

# Division mapping
DIVISIONS = {
    'Rangpur': 'Rangpur Division',
    'Lalmonirhat': 'Rangpur Division',
    'Thakurgaon': 'Rangpur Division',
    'Nilphamari': 'Rangpur Division',
    'Bogura': 'Greater Bogura',
    'Sherpur': 'Greater Bogura',
    'Dupcachia': 'Greater Bogura',
    'Naogaon': 'Greater Bogura',
    'Kushtia': 'Greater Bogura',
    'Jhenaidah': 'Jhenaidah Division',
    'Barishal': 'Jhenaidah Division',
    'Dhaka': 'Dhaka',
    'Chattagram': 'Chattagram'
}


def get_zone(division: str) -> str:
    """Determine zone based on division"""
    if not division:
        return 'Central'
    
    division_lower = division.lower()
    if 'rangpur' in division_lower:
        return 'North'
    elif 'bogura' in division_lower:
        return 'North'
    elif 'jhenaidah' in division_lower:
        return 'South'
    elif 'dhaka' in division_lower:
        return 'Central'
    elif 'chattagram' in division_lower or 'chittagong' in division_lower:
        return 'East'
    else:
        return 'Central'


def extract_month_year_from_filename(filename: str) -> Tuple[Optional[int], Optional[int]]:
    """
    Extract month and year from filename.
    Supports formats like:
    - Sales_Collection_Nov_2025.xlsx
    - Sales_Collection_November_2025.xlsx
    - Sales_11_2025.xlsx
    - 2025_Nov_Sales.xlsx
    """
    filename_lower = filename.lower()
    month = None
    year = None
    
    # Try to find year (4 digits)
    year_match = re.search(r'(20\d{2})', filename)
    if year_match:
        year = int(year_match.group(1))
    
    # Try to find month name
    for month_name, month_num in MONTH_MAP.items():
        if month_name in filename_lower:
            month = month_num
            break
    
    # Try to find month number (01-12 or 1-12)
    if month is None:
        month_match = re.search(r'[_\-\s](\d{1,2})[_\-\s]', filename)
        if month_match:
            potential_month = int(month_match.group(1))
            if 1 <= potential_month <= 12:
                month = potential_month
    
    return month, year


def extract_month_year_from_excel(file_content, sheet_name=0) -> Tuple[Optional[int], Optional[int]]:
    """
    Extract month and year from Excel file header rows.
    Looks for patterns like "November 2025" or "Nov-2025" in first few rows.
    """
    try:
        df_header = pd.read_excel(file_content, sheet_name=sheet_name, nrows=5, header=None)
        file_content.seek(0)  # Reset file pointer
        
        for row in df_header.values:
            for cell in row:
                if pd.notna(cell):
                    cell_str = str(cell).lower()
                    
                    # Try to find year
                    year_match = re.search(r'(20\d{2})', cell_str)
                    year = int(year_match.group(1)) if year_match else None
                    
                    # Try to find month
                    month = None
                    for month_name, month_num in MONTH_MAP.items():
                        if month_name in cell_str:
                            month = month_num
                            break
                    
                    if month and year:
                        return month, year
    except:
        pass
    
    return None, None


def get_or_create_time(db: Session, month: int, year: int) -> DimTime:
    """Get or create time dimension record"""
    time_dim = db.query(DimTime).filter(
        DimTime.month == month,
        DimTime.year == year
    ).first()
    
    if not time_dim:
        current_date = date.today()
        time_dim = DimTime(
            date=date(year, month, 1),
            day=1,
            month=month,
            month_name=get_month_name(month),
            month_short=get_month_short(month),
            quarter=get_quarter(month),
            quarter_name=f"Q{get_quarter(month)}",
            year=year,
            fiscal_year=get_fiscal_year(month, year),
            is_current_month=1 if (month == current_date.month and year == current_date.year) else 0,
            is_current_year=1 if year == current_date.year else 0
        )
        db.add(time_dim)
        db.flush()
    
    return time_dim


def get_or_create_region(db: Session, area_code: str, area_name: str, division: str) -> DimRegion:
    """Get or create region dimension record"""
    region = db.query(DimRegion).filter(DimRegion.area_name == area_name).first()
    
    if not region:
        region = DimRegion(
            area_code=area_code,
            area_name=area_name,
            division=division,
            zone=get_zone(division),
            region_type='Area',
            is_active=1
        )
        db.add(region)
        db.flush()
    
    return region


def get_or_create_product(db: Session, product_name: str, category: str = None) -> DimProduct:
    """Get or create product dimension record"""
    product = db.query(DimProduct).filter(DimProduct.product_name == product_name).first()
    
    if not product:
        product = DimProduct(
            product_name=product_name,
            product_category=category or 'General',
            is_active=1
        )
        db.add(product)
        db.flush()
    
    return product


def delete_existing_sales_data(db: Session, time_id: int) -> int:
    """Delete existing sales data for a specific time period"""
    deleted = db.query(FactSales).filter(FactSales.time_id == time_id).delete()
    db.flush()
    return deleted


def delete_existing_product_data(db: Session, time_id: int) -> int:
    """Delete existing product performance data for a specific time period"""
    deleted = db.query(FactProductPerformance).filter(FactProductPerformance.time_id == time_id).delete()
    db.flush()
    return deleted


def process_sales_collection_file(file_content, filename: str, db: Session, 
                                   override_month: int = None, override_year: int = None) -> Tuple[bool, str, Dict[str, Any]]:
    """
    Process Sales & Collection Excel file into star schema.
    
    Expected Excel Format:
    Row 1-4: Header info (Company name, Period, etc.)
    Row 5+: Data rows with columns:
        A: Area Code (A, B, C, D, E)
        B: Area Name
        C: Sales Target
        D: Gross Sales
        E: Sales Return
        F: Net Sales
        G: (empty or label)
        H: Collection Target
        I: Total Collection
        J-K: (details)
        L: Cash Collection
        M-N: (details)
        O: Credit Collection
        P-Q: (details)
        R: Seed Collection
    """
    try:
        # Extract month/year from filename first, then from Excel content
        month, year = extract_month_year_from_filename(filename)
        
        if month is None or year is None:
            file_content.seek(0)
            excel_month, excel_year = extract_month_year_from_excel(file_content)
            month = month or excel_month
            year = year or excel_year
        
        # Apply overrides if provided
        if override_month:
            month = override_month
        if override_year:
            year = override_year
        
        # Default to current month/year if still not found
        if month is None:
            month = date.today().month
        if year is None:
            year = date.today().year
        
        # Get or create time dimension
        time_dim = get_or_create_time(db, month, year)
        
        # DELETE existing data for this month/year FIRST
        deleted_count = delete_existing_sales_data(db, time_dim.time_id)
        
        records_processed = {
            'month': month,
            'year': year,
            'month_name': get_month_name(month),
            'deleted_records': deleted_count,
            'regions_processed': 0,
            'fact_sales_inserted': 0
        }
        
        # Read the Excel file
        file_content.seek(0)
        df_raw = pd.read_excel(file_content, sheet_name=0, skiprows=4, header=None)
        
        for idx, row in df_raw.iterrows():
            area_code = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
            
            if area_code in ['A', 'B', 'C', 'D', 'E']:
                area_name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                
                if area_name and 'Total' not in area_name and 'total' not in area_name.lower():
                    division = DIVISIONS.get(area_name, 'Unknown')
                    
                    # Get or create region dimension
                    region = get_or_create_region(db, area_code, area_name, division)
                    records_processed['regions_processed'] += 1
                    
                    # Parse sales data (handle NaN as 0)
                    def safe_numeric(val):
                        try:
                            num = pd.to_numeric(val, errors='coerce')
                            return 0 if pd.isna(num) else float(num)
                        except:
                            return 0
                    
                    sales_target = safe_numeric(row.iloc[2])
                    gross_sales = safe_numeric(row.iloc[3])
                    sales_return = safe_numeric(row.iloc[4])
                    net_sales = safe_numeric(row.iloc[5])
                    
                    # Parse collection data
                    coll_target = safe_numeric(row.iloc[7]) if len(row) > 7 else 0
                    total_coll = safe_numeric(row.iloc[8]) if len(row) > 8 else 0
                    cash_coll = safe_numeric(row.iloc[11]) if len(row) > 11 else 0
                    credit_coll = safe_numeric(row.iloc[14]) if len(row) > 14 else 0
                    seed_coll = safe_numeric(row.iloc[17]) if len(row) > 17 else 0
                    
                    # Calculate metrics
                    sales_ach_pct = round((net_sales / sales_target * 100), 2) if sales_target > 0 else 0
                    coll_ach_pct = round((total_coll / coll_target * 100), 2) if coll_target > 0 else 0
                    return_rate = round((sales_return / gross_sales * 100), 2) if gross_sales > 0 else 0
                    outstanding = net_sales - total_coll
                    
                    # Insert new fact record
                    db.add(FactSales(
                        region_id=region.region_id,
                        time_id=time_dim.time_id,
                        sales_target=sales_target,
                        gross_sales=gross_sales,
                        sales_return=sales_return,
                        net_sales=net_sales,
                        sales_achievement_pct=sales_ach_pct,
                        coll_target=coll_target,
                        total_collection=total_coll,
                        cash_collection=cash_coll,
                        credit_collection=credit_coll,
                        seed_collection=seed_coll,
                        coll_achievement_pct=coll_ach_pct,
                        outstanding=outstanding,
                        return_rate_pct=return_rate
                    ))
                    records_processed['fact_sales_inserted'] += 1
        
        db.commit()
        
        message = f"Sales & Collection data for {get_month_name(month)} {year} processed successfully. "
        if deleted_count > 0:
            message += f"Replaced {deleted_count} existing records. "
        message += f"Inserted {records_processed['fact_sales_inserted']} new records."
        
        return True, message, records_processed
        
    except Exception as e:
        db.rollback()
        return False, f"Error processing file: {str(e)}", {}


def process_product_comparison_file(file_content, filename: str, db: Session,
                                     override_month: int = None, override_year: int = None) -> Tuple[bool, str, Dict[str, Any]]:
    """
    Process Product Sales Comparison Excel file into star schema.
    
    Expected Excel Format:
    Sheet 1: "Monthly Value" - Product value comparison
    Sheet 2: "Monthly Volume" - Product volume comparison
    
    Columns:
        A: Index/SL
        B: Product Name
        C: Previous Year Value/Volume
        D: Current Year Value/Volume
        E: (optional)
        F: Growth %
    """
    try:
        # Extract month/year
        month, year = extract_month_year_from_filename(filename)
        
        if month is None or year is None:
            file_content.seek(0)
            excel_month, excel_year = extract_month_year_from_excel(file_content, 'Monthly Value')
            month = month or excel_month
            year = year or excel_year
        
        # Apply overrides
        if override_month:
            month = override_month
        if override_year:
            year = override_year
        
        # Default
        if month is None:
            month = date.today().month
        if year is None:
            year = date.today().year
        
        prev_year = year - 1
        
        # Get or create time dimensions
        time_prev = get_or_create_time(db, month, prev_year)
        time_curr = get_or_create_time(db, month, year)
        
        # DELETE existing data for both years
        deleted_prev = delete_existing_product_data(db, time_prev.time_id)
        deleted_curr = delete_existing_product_data(db, time_curr.time_id)
        
        records_processed = {
            'month': month,
            'year': year,
            'month_name': get_month_name(month),
            'deleted_records': deleted_prev + deleted_curr,
            'products_processed': 0,
            'fact_records_inserted': 0
        }
        
        # Read Value sheet
        file_content.seek(0)
        try:
            df_value = pd.read_excel(file_content, sheet_name='Monthly Value', skiprows=4)
        except:
            file_content.seek(0)
            df_value = pd.read_excel(file_content, sheet_name=0, skiprows=4)
        
        df_value = df_value.dropna(how='all')
        
        # Assign column names based on available columns
        num_cols = len(df_value.columns)
        if num_cols >= 6:
            df_value.columns = ['Index', 'Product_Name', 'Value_Prev', 'Value_Curr', 'Extra', 'Growth'][:num_cols]
        elif num_cols >= 4:
            df_value.columns = ['Index', 'Product_Name', 'Value_Prev', 'Value_Curr'][:num_cols]
        
        # Read Volume sheet
        file_content.seek(0)
        try:
            df_volume = pd.read_excel(file_content, sheet_name='Monthly Volume', skiprows=4)
            df_volume = df_volume.dropna(how='all')
            if len(df_volume.columns) >= 6:
                df_volume.columns = ['Index', 'Product_Name', 'Volume_Prev', 'Volume_Curr', 'Extra', 'Growth'][:len(df_volume.columns)]
            elif len(df_volume.columns) >= 4:
                df_volume.columns = ['Index', 'Product_Name', 'Volume_Prev', 'Volume_Curr'][:len(df_volume.columns)]
        except:
            df_volume = pd.DataFrame()
        
        # Process products
        products = df_value['Product_Name'].dropna().unique()
        exclude_keywords = ['Product Name', 'Surovi', 'Monthly', 'Period', 'Total', 'TOTAL', 'Grand', 'SL', 'No']
        products = [p for p in products if not any(x.lower() in str(p).lower() for x in exclude_keywords)]
        
        def safe_numeric(val):
            try:
                num = pd.to_numeric(val, errors='coerce')
                return 0 if pd.isna(num) else float(num)
            except:
                return 0
        
        for product_name in products:
            product_name = str(product_name).strip()
            if not product_name or len(product_name) < 2:
                continue
                
            # Get or create product dimension
            product = get_or_create_product(db, product_name)
            records_processed['products_processed'] += 1
            
            # Get value data
            value_row = df_value[df_value['Product_Name'] == product_name]
            
            val_prev = 0
            val_curr = 0
            vol_prev = 0
            vol_curr = 0
            
            if not value_row.empty:
                val_prev = safe_numeric(value_row['Value_Prev'].values[0]) if 'Value_Prev' in value_row.columns else 0
                val_curr = safe_numeric(value_row['Value_Curr'].values[0]) if 'Value_Curr' in value_row.columns else 0
            
            # Get volume data if available
            if not df_volume.empty:
                volume_row = df_volume[df_volume['Product_Name'] == product_name]
                if not volume_row.empty:
                    vol_prev = safe_numeric(volume_row['Volume_Prev'].values[0]) if 'Volume_Prev' in volume_row.columns else 0
                    vol_curr = safe_numeric(volume_row['Volume_Curr'].values[0]) if 'Volume_Curr' in volume_row.columns else 0
            
            # Calculate growth
            value_growth = val_curr - val_prev
            volume_growth = vol_curr - vol_prev
            value_growth_pct = round((value_growth / val_prev * 100), 2) if val_prev > 0 else 0
            volume_growth_pct = round((volume_growth / vol_prev * 100), 2) if vol_prev > 0 else 0
            
            # Insert previous year fact
            db.add(FactProductPerformance(
                product_id=product.product_id,
                time_id=time_prev.time_id,
                sales_value=val_prev,
                sales_volume=vol_prev,
                prev_year_value=0,
                prev_year_volume=0,
                value_growth=0,
                volume_growth=0,
                value_growth_pct=0,
                volume_growth_pct=0
            ))
            records_processed['fact_records_inserted'] += 1
            
            # Insert current year fact with YoY comparison
            db.add(FactProductPerformance(
                product_id=product.product_id,
                time_id=time_curr.time_id,
                sales_value=val_curr,
                sales_volume=vol_curr,
                prev_year_value=val_prev,
                prev_year_volume=vol_prev,
                value_growth=value_growth,
                volume_growth=volume_growth,
                value_growth_pct=value_growth_pct,
                volume_growth_pct=volume_growth_pct
            ))
            records_processed['fact_records_inserted'] += 1
        
        db.commit()
        
        message = f"Product comparison data for {get_month_name(month)} {year} processed successfully. "
        if records_processed['deleted_records'] > 0:
            message += f"Replaced {records_processed['deleted_records']} existing records. "
        message += f"Inserted {records_processed['fact_records_inserted']} new records for {records_processed['products_processed']} products."
        
        return True, message, records_processed
        
    except Exception as e:
        db.rollback()
        return False, f"Error processing file: {str(e)}", {}


def detect_file_type(filename: str) -> str:
    """Detect file type based on filename"""
    filename_lower = filename.lower()
    if 'sales' in filename_lower and 'collection' in filename_lower:
        return 'sales_collection'
    elif 'product' in filename_lower or 'comparison' in filename_lower:
        return 'product_comparison'
    elif 'value' in filename_lower or 'volume' in filename_lower:
        return 'product_comparison'
    else:
        return 'unknown'


def generate_sample_template(template_type: str) -> Tuple[BytesIO, str]:
    """Generate a sample Excel template file for download"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from datetime import datetime
    
    # Get current month/year for filename
    now = datetime.now()
    month_name = now.strftime("%b")
    year = now.year
    
    # Style definitions
    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=11)
    col_header_font = Font(bold=True, size=10)
    col_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    col_header_font_white = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    wb = Workbook()
    
    if template_type == "sales_collection":
        ws = wb.active
        ws.title = "Sales Collection"
        
        # Header rows
        ws['A1'] = "SUROVI AGRO INDUSTRIES LTD."
        ws['A1'].font = header_font
        ws.merge_cells('A1:R1')
        
        ws['A2'] = f"Sales & Collection Report - {month_name} {year}"
        ws['A2'].font = sub_header_font
        ws.merge_cells('A2:R2')
        
        ws['A3'] = "(Sample Template - Replace with actual data)"
        ws.merge_cells('A3:R3')
        
        # Column headers (Row 4)
        headers = [
            "Area Code", "Area Name", "Sales Target", "Gross Sales", "Sales Return", 
            "Net Sales", "Label", "Collection Target", "Total Collection", 
            "Detail1", "Detail2", "Cash Collection", "Detail3", "Detail4",
            "Credit Collection", "Detail5", "Detail6", "Seed Collection"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = col_header_font_white
            cell.fill = col_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data rows (Row 5-9)
        sample_data = [
            ["A", "Rangpur", 5000000, 4800000, 50000, 4750000, "Collection", 4500000, 4200000, "-", "-", 2000000, "-", "-", 1500000, "-", "-", 700000],
            ["B", "Bogura", 4000000, 3900000, 40000, 3860000, "Collection", 3800000, 3500000, "-", "-", 1800000, "-", "-", 1200000, "-", "-", 500000],
            ["C", "Dhaka", 6000000, 5800000, 60000, 5740000, "Collection", 5500000, 5200000, "-", "-", 2500000, "-", "-", 2000000, "-", "-", 700000],
            ["D", "Chattogram", 5500000, 5300000, 55000, 5245000, "Collection", 5000000, 4800000, "-", "-", 2300000, "-", "-", 1800000, "-", "-", 700000],
            ["E", "Sylhet", 3500000, 3400000, 35000, 3365000, "Collection", 3200000, 3000000, "-", "-", 1500000, "-", "-", 1000000, "-", "-", 500000],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if isinstance(value, (int, float)) and value != "-":
                    cell.number_format = '#,##0'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        for col in 'CDEFHILORS':
            ws.column_dimensions[col].width = 15
        
        filename = f"Sales_Collection_{month_name}_{year}_TEMPLATE.xlsx"
        
    elif template_type == "product_comparison":
        # Sheet 1: Monthly Value
        ws1 = wb.active
        ws1.title = "Monthly Value"
        
        # Header rows
        ws1['A1'] = "SUROVI AGRO INDUSTRIES LTD."
        ws1['A1'].font = header_font
        ws1.merge_cells('A1:F1')
        
        ws1['A2'] = f"Product Comparison (Value) - {month_name} {year}"
        ws1['A2'].font = sub_header_font
        ws1.merge_cells('A2:F2')
        
        ws1['A3'] = "(Sample Template - Replace with actual data)"
        ws1.merge_cells('A3:F3')
        
        # Column headers
        value_headers = ["SL", "Product Name", f"{month_name} {year-1}", f"{month_name} {year}", "Diff", "Growth %"]
        for col, header in enumerate(value_headers, 1):
            cell = ws1.cell(row=4, column=col, value=header)
            cell.font = col_header_font_white
            cell.fill = col_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Sample product data
        products_value = [
            [1, "Surovi Ghee 200ml", 1500000, 1800000, 300000, "20%"],
            [2, "Surovi Ghee 500ml", 2500000, 2800000, 300000, "12%"],
            [3, "Surovi Ghee 1000ml", 3000000, 3500000, 500000, "17%"],
            [4, "Surovi Butter 100g", 800000, 950000, 150000, "19%"],
            [5, "Surovi Butter 200g", 1200000, 1400000, 200000, "17%"],
        ]
        
        for row_idx, row_data in enumerate(products_value, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if isinstance(value, (int, float)) and col_idx in [3, 4, 5]:
                    cell.number_format = '#,##0'
        
        ws1.column_dimensions['B'].width = 25
        for col in 'CDEF':
            ws1.column_dimensions[col].width = 15
        
        # Sheet 2: Monthly Volume
        ws2 = wb.create_sheet("Monthly Volume")
        
        # Header rows
        ws2['A1'] = "SUROVI AGRO INDUSTRIES LTD."
        ws2['A1'].font = header_font
        ws2.merge_cells('A1:F1')
        
        ws2['A2'] = f"Product Comparison (Volume) - {month_name} {year}"
        ws2['A2'].font = sub_header_font
        ws2.merge_cells('A2:F2')
        
        ws2['A3'] = "(Sample Template - Replace with actual data)"
        ws2.merge_cells('A3:F3')
        
        # Column headers
        volume_headers = ["SL", "Product Name", f"{month_name} {year-1}", f"{month_name} {year}", "Diff", "Growth %"]
        for col, header in enumerate(volume_headers, 1):
            cell = ws2.cell(row=4, column=col, value=header)
            cell.font = col_header_font_white
            cell.fill = col_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Sample volume data
        products_volume = [
            [1, "Surovi Ghee 200ml", 5000, 6000, 1000, "20%"],
            [2, "Surovi Ghee 500ml", 4000, 4500, 500, "13%"],
            [3, "Surovi Ghee 1000ml", 2500, 3000, 500, "20%"],
            [4, "Surovi Butter 100g", 8000, 9500, 1500, "19%"],
            [5, "Surovi Butter 200g", 6000, 7000, 1000, "17%"],
        ]
        
        for row_idx, row_data in enumerate(products_volume, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if isinstance(value, (int, float)) and col_idx in [3, 4, 5]:
                    cell.number_format = '#,##0'
        
        ws2.column_dimensions['B'].width = 25
        for col in 'CDEF':
            ws2.column_dimensions[col].width = 15
        
        filename = f"Product_Comparison_{month_name}_{year}_TEMPLATE.xlsx"
    
    else:
        raise ValueError(f"Unknown template type: {template_type}")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, filename


def get_sample_format_info() -> Dict[str, Any]:
    """Return sample format information for file uploads"""
    return {
        'sales_collection': {
            'description': 'Sales & Collection Monthly Data',
            'filename_format': 'Sales_Collection_<Month>_<Year>.xlsx (e.g., Sales_Collection_Nov_2025.xlsx)',
            'columns': [
                {'col': 'A', 'name': 'Area Code', 'example': 'A, B, C, D, E'},
                {'col': 'B', 'name': 'Area Name', 'example': 'Rangpur, Bogura, Dhaka'},
                {'col': 'C', 'name': 'Sales Target', 'example': '5000000'},
                {'col': 'D', 'name': 'Gross Sales', 'example': '4800000'},
                {'col': 'E', 'name': 'Sales Return', 'example': '50000'},
                {'col': 'F', 'name': 'Net Sales', 'example': '4750000'},
                {'col': 'G', 'name': '(Label)', 'example': 'Collection'},
                {'col': 'H', 'name': 'Collection Target', 'example': '4500000'},
                {'col': 'I', 'name': 'Total Collection', 'example': '4200000'},
                {'col': 'J-K', 'name': '(Details)', 'example': '-'},
                {'col': 'L', 'name': 'Cash Collection', 'example': '2000000'},
                {'col': 'M-N', 'name': '(Details)', 'example': '-'},
                {'col': 'O', 'name': 'Credit Collection', 'example': '1500000'},
                {'col': 'P-Q', 'name': '(Details)', 'example': '-'},
                {'col': 'R', 'name': 'Seed Collection', 'example': '700000'},
            ],
            'notes': [
                'First 4 rows are header (Company name, Period info)',
                'Data starts from row 5',
                'Include month/year in filename OR in header rows',
                'Existing data for the same month will be replaced'
            ]
        },
        'product_comparison': {
            'description': 'Product Sales Comparison (YoY)',
            'filename_format': 'Product_Comparison_<Month>_<Year>.xlsx (e.g., Product_Comparison_Nov_2025.xlsx)',
            'sheets': [
                {
                    'name': 'Monthly Value',
                    'columns': [
                        {'col': 'A', 'name': 'SL/Index', 'example': '1, 2, 3'},
                        {'col': 'B', 'name': 'Product Name', 'example': 'Surovi Ghee 200ml'},
                        {'col': 'C', 'name': 'Previous Year Value', 'example': '1500000'},
                        {'col': 'D', 'name': 'Current Year Value', 'example': '1800000'},
                        {'col': 'E', 'name': '(Optional)', 'example': '-'},
                        {'col': 'F', 'name': 'Growth %', 'example': '20%'},
                    ]
                },
                {
                    'name': 'Monthly Volume',
                    'columns': [
                        {'col': 'A', 'name': 'SL/Index', 'example': '1, 2, 3'},
                        {'col': 'B', 'name': 'Product Name', 'example': 'Surovi Ghee 200ml'},
                        {'col': 'C', 'name': 'Previous Year Volume', 'example': '5000'},
                        {'col': 'D', 'name': 'Current Year Volume', 'example': '6000'},
                        {'col': 'E', 'name': '(Optional)', 'example': '-'},
                        {'col': 'F', 'name': 'Growth %', 'example': '20%'},
                    ]
                }
            ],
            'notes': [
                'Two sheets required: "Monthly Value" and "Monthly Volume"',
                'First 4 rows are header',
                'Data starts from row 5',
                'Product names must match across both sheets',
                'Existing data for the same month will be replaced'
            ]
        }
    }
